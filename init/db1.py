import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import NameDB1
import pymysql.cursors


def check_db():
    try:
        conn = pymysql.connect(host="localhost",
                               user="root",
                               password="root",
                               database=NameDB1.namedb.get_name_db(),
                               cursorclass=pymysql.cursors.Cursor)
        print("Вы подключились")
    except pymysql.err.MySQLError:
        conn = pymysql.connect(host="localhost",
                               user="root",
                               password="root",
                               cursorclass=pymysql.cursors.Cursor)

        cursor = conn.cursor()
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {NameDB1.namedb.get_name_db()}")
        print("Вы создали БД")
    return conn


def con_db():
    return pymysql.connect(host="localhost",
                           user="root",
                           password="root",
                           database=NameDB1.namedb.get_name_db(),
                           cursorclass=pymysql.cursors.Cursor)


def check_table():
    connection = con_db()
    cursor = connection.cursor()
    try:
        cursor.execute(f"SELECT * FROM {NameDB1.namedb.get_name_db()}")
        print("Таблица подключена")
    except pymysql.err.MySQLError:
        cursor.execute(
            f"CREATE TABLE IF NOT EXISTS {NameDB1.namedb.get_name_tb()} (id bigint NOT NULL AUTO_INCREMENT,player_X varchar(100) NOT NULL,player_O varchar(100) NOT NULL, history_x varchar(1000) NOT NULL, history_o varchar(1000) NOT NULL, PRIMARY KEY (id)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci;")
    connection.commit()
    print("Таблица создана и подключена")
    return


def save_result(x, o, arr_x, arr_o):
    connection = con_db()
    try:
        cursor = connection.cursor()
        cursor.execute(f"INSERT INTO " + NameDB1.namedb.get_name_tb() + f" (player_X, player_O, history_x, history_o) VALUES (%s, %s, %s, %s)",
                       (str(x), str(o), str(arr_x), str(arr_o)))
        connection.commit()

        cursor = connection.cursor()
        cursor.execute(f"SELECT * FROM {NameDB1.namedb.get_name_tb()}")
        print(cursor.fetchall()[-1])
    except pymysql.err.DataError as e:
        print('Ошибка с данными:', e)
    except pymysql.err.DatabaseError as e:
        print(e)
    return


def save_to_excel():
    connection = con_db()
    try:
        new_df = pd.read_sql("SELECT * FROM " + NameDB1.namedb.get_name_tb(), connection)
        wb = openpyxl.Workbook()
        ws = wb.active

        for r in dataframe_to_rows(new_df, index=False, header=True):
            ws.append(r)

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        file1 = input("Введите имя файла с расширением xlsx: ")
        wb.save(file1)
        print(new_df)

    except pymysql.err.DatabaseError as e:
        print(e)
    return
